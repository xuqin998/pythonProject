from openpyxl import load_workbook

class HandlerExcel:
    def __init__(self,filename,sheetname):
        self.wb=load_workbook(filename=filename)
        self.sheet=self.wb[sheetname]

    def get_data(self):
        case_list=[]
        cases=list(self.sheet.iter_rows(values_only=True))
        for case in cases[1:]:
            case_dict=dict(list(zip(cases[0],case)))
            case_list.append(case_dict)
        self.close_wb()
        return case_list

    def close_wb(self):
        self.wb.close()



